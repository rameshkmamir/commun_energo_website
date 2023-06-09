import json
from urllib.parse import unquote_plus

from django.db.models import Q
from django.shortcuts import render, redirect

from messages_home.models import Conversation
from django.contrib.auth.models import User
from collections import defaultdict
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

def index(request):
    user = request.user
    if user.is_authenticated and user.groups.filter(name='Обычные пользователи').exists():
        return redirect('conversations_list')
    if user.is_authenticated and user.groups.filter(name='Поддержка').exists():
        conversations = Conversation.objects.filter(Q(user2__username__icontains=user))
    else:
        conversations = Conversation.objects.all()

    data = {
        'labels': [],
        'datasets': [
            {'label': 'Активная', 'data': []},
            {'label': 'Закрыта', 'data': []},
            {'label': 'Отложена', 'data': []}
        ]
    }

    recipient = request.GET.get('user_name')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    context = {}
    if recipient:
        recipient = unquote_plus(recipient)
        recipient_parts = recipient.split()
        if len(recipient_parts) >= 2:
            recipient_filter = Q(user2__first_name__icontains=recipient_parts[0]) | Q(
                user2__last_name__icontains=recipient_parts[1])
        else:
            recipient_filter = Q(user2__first_name__icontains=recipient) | Q(
                user2__last_name__icontains=recipient)
        conversations = conversations.filter(recipient_filter)
        context['recipient'] = recipient
    if date_start and date_end:
        date_filter = Q(created_at__gte=date_start) & Q(created_at__lte=date_end)
        conversations = conversations.filter(date_filter)
        context['date_start'] = date_start
        context['date_end'] = date_end
    if date_start and not (date_end):
        date_filter = Q(created_at__gte=date_start)
        conversations = conversations.filter(date_filter)
        context['date_start'] = date_start
    if date_end and not (date_start):
        date_filter = Q(created_at__lte=date_end)
        conversations = conversations.filter(date_filter)
        context['date_end'] = date_end

    for conversation in conversations:
        created_at = conversation.created_at.date().strftime('%Y-%m-%d')
        data['labels'].append(created_at)
        if conversation.status == 'Активная':
            data['datasets'][0]['data'].append(created_at)
        elif conversation.status == 'Закрыта':
            data['datasets'][1]['data'].append(created_at)
        elif conversation.status == 'Отложена':
            data['datasets'][2]['data'].append(created_at)

    context.update({
        'data': json.dumps(data),
    })
    return render(request, "main/index.html", context=context)


def settings(request):
    user = request.user

    data = {
        "title": "Настройки",
    }
    if user.is_authenticated and user.groups.filter(name='Администратор').exists():
        return render(request, "main/settings.html", data)
    else:
        previous_page = request.META.get('HTTP_REFERER')
        return redirect(previous_page)

def report(request):
    pdfmetrics.registerFont(TTFont('Roboto', 'main/templates/main/Roboto/Roboto-Regular.ttf'))
    style = [('FONTNAME', (0, 0), (-1, -1), 'Roboto')]
    users = User.objects.filter(groups__name__in=['Администратор', 'Поддержка'])

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    doc = SimpleDocTemplate(response, pagesize=letter, style=style)

    elements = []

    for user in users:
        conversations = Conversation.objects.filter(user2=user)

        # Создание словаря для хранения данных по датам и статусам
        data_dict = defaultdict(lambda: [0, 0, 0])

        for conversation in conversations:
            date = conversation.created_at.date()
            status = conversation.status

            # Увеличение количества заявок по статусу для соответствующей даты
            if status == 'Активная':
                data_dict[date][0] += 1
            elif status == 'Отложена':
                data_dict[date][1] += 1
            elif status == 'Закрыта':
                data_dict[date][2] += 1

        # Создание таблицы данных
        table_data = [[f'Отчет для сотрудника : {user.first_name} {user.last_name}']]
        table_data.append(['Дата', 'Активная', 'Отложена', 'Закрыта'])

        for date, counts in data_dict.items():
            row = [date, counts[0], counts[1], counts[2]]
            table_data.append(row)

        # Определение ширины столбцов
        table_width = doc.width
        column_widths = [table_width * 0.25, table_width * 0.25, table_width * 0.25, table_width * 0.25]

        # Создание таблицы для пользователя
        table = Table(table_data, colWidths=column_widths, style=style)
        style = TableStyle([
          ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
          ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
          ('GRID', (0, 0), (-1, -1), 1, colors.black),
          ('SPAN', (0, 0), (-1, 0)),
          ('FONTNAME', (0, 0), (-1, -1), 'Roboto'),
        ])
        table.setStyle(style)

        elements.append(table)

    doc.build(elements)

    return response


def report_xlsx(request):
  users = User.objects.filter(groups__name__in=['Администратор', 'Поддержка'])

  workbook = openpyxl.Workbook()
  sheet = workbook.active

  row_num = 1

  for user in users:
    conversations = Conversation.objects.filter(user2=user)

    data_dict = defaultdict(lambda: [0, 0, 0])  # Обнуляем словарь для каждого пользователя

    for conversation in conversations:
      date = conversation.created_at.date()
      status = conversation.status

      if status == 'Активная':
        data_dict[date][0] += 1
      elif status == 'Отложена':
        data_dict[date][1] += 1
      elif status == 'Закрыта':
        data_dict[date][2] += 1

    sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    cell = sheet.cell(row=row_num, column=1, value=f'Отчет для сотрудника: {user.first_name} {user.last_name}')
    cell.alignment = Alignment(horizontal='left')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    row_num += 1

    headers = ['Дата', 'Активная', 'Отложена', 'Закрыта']
    for col_num, header in enumerate(headers, 1):
      col_letter = get_column_letter(col_num)
      cell = sheet.cell(row=row_num, column=col_num, value=header)
      cell.alignment = Alignment(horizontal='left')
      cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                           bottom=Side(style='thin'))
      sheet.column_dimensions[col_letter].width = 12

    row_num += 1

    for date, counts in data_dict.items():
      row = [date.strftime('%Y-%m-%d'), counts[0], counts[1], counts[2]]
      for col_num, value in enumerate(row, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(horizontal='left')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
      row_num += 1

    row_num += 1

  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
  workbook.save(response)

  return response